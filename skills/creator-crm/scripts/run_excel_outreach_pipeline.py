#!/usr/bin/env python3
"""Run Creator CRM from a Kalodata Excel attachment.

Flow:
1. Parse Excel creator ranking export.
2. Upsert/dedupe against RDS creator pool.
3. For creators without complete analysis, fetch 12 covers, generate grid, upload grids to OSS.
4. Run video scoring + vibe + category tagging, then entry-screen rules.
5. Generate batch outreach copy and write only key outreach rows to an output Feishu table.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import requests


SKILL_DIR = Path(__file__).resolve().parents[1]
WORKSPACE_DIR = SKILL_DIR.parents[1]
PROFILE_CARD_DIR = WORKSPACE_DIR / "skills" / "creator-profile-card"
REPO_ROOT = SKILL_DIR.parents[1]

for path in (REPO_ROOT, SKILL_DIR, PROFILE_CARD_DIR):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from workspace_support import load_repo_env

load_repo_env()

from core.creator_repository import CreatorRepository  # noqa: E402
from core.feishu_reader import CreatorRecord, FeishuBitableReader  # noqa: E402
from core.kalodata_fetcher import KalodataAccessBlocked  # noqa: E402
from core.llm_analyzer import CategoryTaggingAgent, CombinedScoringVibeAgent, EntryScreenAgent  # noqa: E402
from core.oss_assets import CreatorAssetStorage  # noqa: E402
from core.sub_agents import GridGeneratorAgent, VideoFetcherAgent  # noqa: E402
from run_pipeline import apply_reusable_analysis, should_skip_creator_pool_relationship  # noqa: E402
from scripts.generate_entry_outreach_drafts import (  # noqa: E402
    as_float,
    as_text,
    build_profile_context,
    generate_batch_template,
    render_creator_template,
)


OUT_DIR = SKILL_DIR / "output" / "excel_outreach_runs"
GRID_DIR = SKILL_DIR / "output" / "grids" / "excel_outreach"


FINAL_FIELDS: List[Dict[str, Any]] = [
    {"field_name": "建联批次号", "type": 1, "ui_type": "Text"},
    {"field_name": "达人handle", "type": 1, "ui_type": "Text"},
    {"field_name": "达人名称", "type": 1, "ui_type": "Text"},
    {"field_name": "TikTok链接", "type": 15, "ui_type": "Url"},
    {"field_name": "粉丝数", "type": 2, "ui_type": "Number"},
    {"field_name": "榜单成交金额", "type": 2, "ui_type": "Number"},
    {"field_name": "准入评分", "type": 2, "ui_type": "Number"},
    {"field_name": "达人分层", "type": 1, "ui_type": "Text"},
    {"field_name": "适配类目", "type": 1, "ui_type": "Text"},
    {"field_name": "计划建联产品", "type": 1, "ui_type": "Text"},
    {"field_name": "批量建联话术", "type": 1, "ui_type": "Text"},
    {"field_name": "批量建联话术本地语言", "type": 1, "ui_type": "Text"},
    {"field_name": "建联发送状态", "type": 1, "ui_type": "Text"},
    {"field_name": "达人已回复", "type": 7, "ui_type": "Checkbox"},
    {"field_name": "回复备注", "type": 1, "ui_type": "Text"},
    {"field_name": "进入维护状态", "type": 1, "ui_type": "Text"},
    {"field_name": "RDS达人UID", "type": 1, "ui_type": "Text"},
]


def normalize_handle(value: Any) -> str:
    text = as_text(value)
    if "tiktok.com/@" in text:
        text = text.split("tiktok.com/@", 1)[1]
    text = text.split("?", 1)[0].strip().strip("/")
    return text[1:] if text.startswith("@") else text


def safe_number(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_excel_rows(excel_path: str, *, sheet_name: str = "LIST_CREATOR", limit: int = 0, offset: int = 0) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("需要 openpyxl 才能读取 Excel。请使用 Codex bundled Python 或安装 openpyxl。") from exc

    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    target_sheet = sheet_name if sheet_name in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[target_sheet]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]

    records: List[Dict[str, Any]] = []
    for row_index, row in enumerate(rows, 2):
        raw = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
        handle = normalize_handle(raw.get("达人handle（达人账号）") or raw.get("达人账号") or raw.get("TikTok链接"))
        kalodata_url = as_text(raw.get("Kalodata详情页链接") or raw.get("Kalodata链接"))
        if not handle or not kalodata_url:
            continue
        tk_url = as_text(raw.get("TikTok链接")) or f"https://www.tiktok.com/@{handle}"
        records.append({
            "record_id": f"excel_row_{row_index}",
            "tk_handle": handle,
            "tk_url": tk_url,
            "kalodata_url": kalodata_url,
            "creator_name": as_text(raw.get("达人名称")),
            "followers_count": safe_number(raw.get("粉丝数")),
            "gmv": safe_number(raw.get("成交金额(¥)")),
            "video_count": safe_number(raw.get("视频数量")),
            "video_gmv": safe_number(raw.get("视频金额(¥)")),
            "content_views": safe_number(raw.get("内容观看")),
            "raw_fields": raw,
        })

    if offset:
        records = records[offset:]
    if limit:
        records = records[:limit]
    return records


def get_access_token() -> str:
    return FeishuBitableReader()._get_access_token()


def parse_feishu_url(feishu_url: str) -> Tuple[str, str]:
    table_match = re.search(r"[?&]table=([^&]+)", feishu_url)
    if not table_match:
        raise ValueError(f"飞书 URL 缺少 table 参数: {feishu_url}")
    table_id = table_match.group(1)

    base_match = re.search(r"/base/([a-zA-Z0-9]+)", feishu_url)
    if base_match:
        return base_match.group(1), table_id

    wiki_match = re.search(r"/wiki/([a-zA-Z0-9]+)", feishu_url)
    if not wiki_match:
        raise ValueError(f"无法解析飞书 URL: {feishu_url}")

    response = requests.get(
        "https://open.feishu.cn/open-apis/wiki/v2/spaces/get_node",
        headers={"Authorization": f"Bearer {get_access_token()}"},
        params={"token": wiki_match.group(1)},
        timeout=30,
    )
    result = response.json()
    if result.get("code") != 0:
        raise RuntimeError(f"解析 wiki token 失败: {json.dumps(result, ensure_ascii=False)}")
    return result["data"]["node"]["obj_token"], table_id


def list_fields(app_token: str, table_id: str, token: str) -> set:
    url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/fields"
    names = set()
    page_token = None
    while True:
        params = {"page_size": 100}
        if page_token:
            params["page_token"] = page_token
        response = requests.get(url, headers={"Authorization": f"Bearer {token}"}, params=params, timeout=30)
        result = response.json()
        if result.get("code") != 0:
            raise RuntimeError(f"读取飞书字段失败: {json.dumps(result, ensure_ascii=False)}")
        for item in result.get("data", {}).get("items", []):
            names.add(item.get("field_name"))
        if not result.get("data", {}).get("has_more"):
            break
        page_token = result.get("data", {}).get("page_token")
    return names


def ensure_final_fields(app_token: str, table_id: str) -> None:
    token = get_access_token()
    existing = list_fields(app_token, table_id, token)
    for spec in FINAL_FIELDS:
        if spec["field_name"] in existing:
            continue
        payload = {"field_name": spec["field_name"], "type": spec["type"], "ui_type": spec["ui_type"]}
        response = requests.post(
            f"https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/fields",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json=payload,
            timeout=30,
        )
        result = response.json()
        if result.get("code") != 0:
            raise RuntimeError(f"创建字段失败 {spec['field_name']}: {json.dumps(result, ensure_ascii=False)}")


def create_feishu_records(app_token: str, table_id: str, rows: List[Dict[str, Any]], batch_size: int = 100) -> int:
    if not rows:
        return 0
    token = get_access_token()
    url = f"https://open.feishu.cn/open-apis/bitable/v1/apps/{app_token}/tables/{table_id}/records/batch_create"
    created = 0
    for start in range(0, len(rows), batch_size):
        batch = [{"fields": row} for row in rows[start:start + batch_size]]
        response = requests.post(
            url,
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json={"records": batch},
            timeout=60,
        )
        result = response.json()
        if result.get("code") != 0:
            raise RuntimeError(f"写入最终飞书表失败: {json.dumps(result, ensure_ascii=False)}")
        created += len(batch)
    return created


def safe_int_score(value: Any, default: int = 3, min_value: int = 1, max_value: int = 5) -> int:
    try:
        output = int(float(value))
        return min(max_value, max(min_value, output))
    except (TypeError, ValueError):
        return default


def safe_float_score(value: Any, default: float = 3.0) -> float:
    try:
        output = float(value)
        return round(min(5.0, max(1.0, output)), 1)
    except (TypeError, ValueError):
        return default


def compact_scoring_result(combined: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "analysis_reason": combined.get("analysis_reason", ""),
        "score_traffic": safe_int_score(combined.get("score_traffic")),
        "score_presence": safe_int_score(combined.get("score_presence")),
        "score_consistency": safe_int_score(combined.get("score_consistency")),
        "score_lighting": safe_int_score(combined.get("score_lighting")),
        "score_background": safe_int_score(combined.get("score_background")),
        "total_score": safe_int_score(combined.get("total_score"), 15, 5, 25),
        "final_star_rating": safe_float_score(combined.get("final_star_rating")),
    }


def build_creator_record(row: Dict[str, Any], entry_result: Dict[str, Any]) -> CreatorRecord:
    fields = dict(row.get("raw_fields") or {})
    fields.update({
        "准入评分": entry_result.get("entry_score"),
        "适配类目": entry_result.get("fit_category"),
        "达人擅长内容形式": entry_result.get("creator_content_mode"),
        "内容类型": entry_result.get("primary_content_type"),
    })
    return CreatorRecord(
        record_id=row["record_id"],
        tk_handle=row["tk_handle"],
        tk_url=row["tk_url"],
        kalodata_url=row["kalodata_url"],
        video_screenshots=None,
        raw_fields=fields,
    )


def generate_batch_outreach(rows: List[Dict[str, Any]], *, product_name: str, product_category: str, selling_points: str, version: str) -> Dict[str, Any]:
    if not rows:
        return {}
    representative = build_creator_record(rows[0], rows[0]["entry_result"])
    relationship = {"history_relation": "陌生", "relationship_stage": "冷", "current_action": "主动新品邀约"}
    product_context = (product_name, product_category, selling_points)
    return generate_batch_template(
        representative,
        version=version,
        product_context=product_context,
        relationship=relationship,
    )


def build_output_row(
    row: Dict[str, Any],
    *,
    batch_id: str,
    product_name: str,
    batch_message: Dict[str, Any],
    include_oss_links: bool,
) -> Dict[str, Any]:
    scoring = row.get("scoring_result") or {}
    category = row.get("category_result") or {}
    vibe = row.get("vibe_result") or {}
    entry = row.get("entry_result") or {}
    draft_cn = strip_batch_greeting(
        render_creator_template(as_text(batch_message.get("message_cn_for_operator")), build_creator_record(row, entry))
    )
    draft_local = strip_batch_greeting(
        render_creator_template(as_text(batch_message.get("message_local")), build_creator_record(row, entry))
    )
    oss_links = "\n".join(ref.get("public_url") or ref.get("object_key") or "" for ref in row.get("oss_grid_refs") or [])

    output = {
        "建联批次号": batch_id,
        "达人handle": row.get("tk_handle"),
        "达人名称": row.get("creator_name"),
        "TikTok链接": {"link": row.get("tk_url"), "text": row.get("tk_url")},
        "粉丝数": row.get("followers_count"),
        "榜单成交金额": row.get("gmv"),
        "准入评分": entry.get("entry_score"),
        "达人分层": entry.get("entry_tier"),
        "适配类目": entry.get("fit_category"),
        "计划建联产品": product_name,
        "批量建联话术": draft_cn,
        "批量建联话术本地语言": draft_local,
        "建联发送状态": "待发送",
        "达人已回复": False,
        "回复备注": "",
        "进入维护状态": "待回复",
        "RDS达人UID": row.get("rds_creator_uid"),
    }
    if include_oss_links:
        output["素材OSS链接"] = oss_links
    return {key: value for key, value in output.items() if value not in (None, "")}


def strip_batch_greeting(message: str) -> str:
    """Remove batch-message leading greetings and creator names.

    Batch outreach copy should be reusable. Keep product/value content, remove
    personalized openers such as "哈喽@name～", "Hello name," or Thai greetings.
    """
    text = (message or "").strip()
    if not text:
        return ""

    patterns = [
        r"^(?:哈喽|你好|您好|嗨|Hi|Hello|Hey)\s*[@＠]?[A-Za-z0-9_.\-\u4e00-\u9fff]+[～~，,、!\s-]*",
        r"^(?:สวัสดีค่ะ|สวัสดีครับ|สวัสดี|หวัดดีค่ะ|หวัดดีครับ)\s*[@＠]?[A-Za-z0-9_.\-]+[～~，,、!\s-]*",
        r"^[@＠][A-Za-z0-9_.\-]+[～~，,、!\s-]*",
    ]
    for pattern in patterns:
        text = re.sub(pattern, "", text, count=1, flags=re.IGNORECASE).strip()

    return text


def process_one(
    row: Dict[str, Any],
    *,
    repository: Optional[CreatorRepository],
    asset_storage: CreatorAssetStorage,
    video_fetcher: VideoFetcherAgent,
    grid_generator: GridGeneratorAgent,
    scoring_agent: CombinedScoringVibeAgent,
    category_agent: CategoryTaggingAgent,
    entry_agent: EntryScreenAgent,
    max_videos: int,
    max_grids: int,
    retention_days: int,
) -> Dict[str, Any]:
    creator_uid = ""
    if repository:
        creator_uid = repository.upsert_creator_from_payload(row, source_status="excel_attachment")
        row["rds_creator_uid"] = creator_uid
        relationship = repository.get_relationship(creator_uid)
        skip_relationship, reason = should_skip_creator_pool_relationship(relationship)
        if skip_relationship:
            row["skipped"] = True
            row["skip_reason"] = reason
            return row
        analysis = repository.get_analysis(creator_uid)
        if analysis and analysis.get("analysis_status") == "complete":
            apply_reusable_analysis(row, analysis)
            row["scoring_result"] = {
                "final_star_rating": row.get("video_final_score"),
                "analysis_reason": analysis.get("score_reason") or "",
            }
            row["category_result"] = {
                "main_category_1": row.get("main_category"),
                "sub_category_1": row.get("sub_category"),
                "analysis_reason": analysis.get("tag_reason") or "",
            }
            row["vibe_result"] = {
                "vibe_tag": analysis.get("vibe_tag") or "",
                "vibe_reason": analysis.get("vibe_reason") or "",
            }
            row["entry_result"] = entry_agent.execute({
                "tk_handle": row["tk_handle"],
                "video_final_score": row.get("video_final_score"),
                "main_category": row.get("main_category"),
                "sub_category": row.get("sub_category"),
            })
            return row

    video_result = video_fetcher.execute({
        "tk_handle": row["tk_handle"],
        "kalodata_url": row["kalodata_url"],
        "max_videos": max_videos,
    })
    cover_urls = (video_result.get("cover_urls") or [])[:max_videos]
    views_list = (video_result.get("views_list") or [])[:max_videos]
    revenue_list = (video_result.get("revenue_list") or [])[:max_videos]
    if len(cover_urls) < 6:
        raise RuntimeError(f"封面数量不足: {len(cover_urls)}/6")

    grid_result = grid_generator.execute({
        "tk_handle": row["tk_handle"],
        "cover_urls": cover_urls,
        "views_list": views_list,
        "revenue_list": revenue_list,
        "output_dir": str(GRID_DIR),
        "max_grids": max_grids,
    })
    grid_paths = grid_result.get("grid_paths", [grid_result["grid_path"]])
    row["grid_paths"] = grid_paths

    oss_refs = []
    for grid_path in grid_paths:
        upload = asset_storage.upload_grid(Path(grid_path), creator_uid=creator_uid, tk_handle=row["tk_handle"], run_id=row["record_id"])
        ref = upload.to_dict()
        if repository:
            ref["asset_id"] = repository.record_asset(
                creator_uid=creator_uid,
                asset_type="grid_image",
                storage_provider=upload.provider,
                bucket=upload.bucket,
                object_key=upload.object_key,
                public_url=upload.public_url,
                source_path=str(grid_path),
                file_name=upload.file_name,
                file_size=upload.file_size,
                file_hash=upload.file_hash,
                retention_days=retention_days,
                meta={"tk_handle": row["tk_handle"], "source": "excel_outreach"},
            )
        oss_refs.append(ref)
    row["oss_grid_refs"] = oss_refs

    combined = scoring_agent.execute({
        "tk_handle": row["tk_handle"],
        "grid_paths": grid_paths,
        "views_list": views_list,
    })
    scoring_result = compact_scoring_result(combined)
    vibe_result = {
        "vibe_tag": combined.get("vibe_tag", "Unknown"),
        "vibe_reason": combined.get("vibe_reason", ""),
    }
    category_result = category_agent.execute({"tk_handle": row["tk_handle"], "grid_paths": grid_paths})
    entry_result = entry_agent.execute({
        "tk_handle": row["tk_handle"],
        "video_final_score": scoring_result.get("final_star_rating"),
        "main_category": category_result.get("main_category_1"),
        "sub_category": category_result.get("sub_category_1"),
    })

    row["scoring_result"] = scoring_result
    row["vibe_result"] = vibe_result
    row["category_result"] = category_result
    row["entry_result"] = entry_result

    if repository and creator_uid:
        repository.upsert_analysis_from_result(
            creator_uid,
            scoring_result=scoring_result,
            category_result=category_result,
            vibe_result=vibe_result,
            screenshot_refs=oss_refs,
            sample_video_refs=[
                {
                    "video_id": video_id,
                    "cover_url": cover_urls[idx] if idx < len(cover_urls) else "",
                    "views": views_list[idx] if idx < len(views_list) else None,
                    "revenue": revenue_list[idx] if idx < len(revenue_list) else None,
                }
                for idx, video_id in enumerate((video_result.get("video_ids") or [])[:max_videos])
            ],
            analysis_status="complete",
        )
    return row


def main() -> int:
    parser = argparse.ArgumentParser(description="Excel 附件驱动 Creator CRM 建联轻筛流程")
    parser.add_argument("--excel", required=True, help="Kalodata 导出的 Excel 文件路径")
    parser.add_argument("--sheet-name", default="LIST_CREATOR")
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--offset", type=int, default=0)
    parser.add_argument("--max-videos", type=int, default=12, help="默认只抓 12 个视频封面")
    parser.add_argument("--max-grids", type=int, default=1)
    parser.add_argument("--batch-id", default=datetime.now().strftime("CRM%Y%m%d%H%M"))
    parser.add_argument("--product-name", default="本轮计划建联产品")
    parser.add_argument("--product-category", default="轻上装")
    parser.add_argument("--selling-points", default="日常好搭、适合低成本短视频展示")
    parser.add_argument("--message-version", default="excel_attachment_batch_v1")
    parser.add_argument("--output-feishu-url", default="", help="最终运营输出表 URL；不传则只输出 JSONL")
    parser.add_argument("--write-feishu", action="store_true", help="写入最终飞书表")
    parser.add_argument("--include-manual", action="store_true", help="最终名单包含 人工查看")
    parser.add_argument("--include-oss-links", action="store_true", help="最终飞书表也写素材 OSS 链接；默认不写过程资产")
    parser.add_argument("--creator-db-url", default="", help="默认读取 CREATOR_CRM_DATABASE_URL 或 LIKEU_AI_DATABASE_URL")
    parser.add_argument("--no-rds", action="store_true")
    parser.add_argument("--continue-on-kalodata-block", action="store_true", help="Kalodata 403 时继续处理后续达人；默认熔断整批")
    parser.add_argument("--retention-days", type=int, default=int(os.environ.get("CREATOR_CRM_ASSET_RETENTION_DAYS", "30") or 30))
    parser.add_argument("--dry-run", action="store_true", help="只解析 Excel 和展示候选，不抓取/打标/写飞书")
    args = parser.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    GRID_DIR.mkdir(parents=True, exist_ok=True)
    rows = parse_excel_rows(args.excel, sheet_name=args.sheet_name, limit=args.limit, offset=args.offset)
    print(f"📖 Excel 有效候选: {len(rows)}")
    if args.dry_run:
        for row in rows[:10]:
            print(f"  - @{row['tk_handle']} gmv={row.get('gmv')} kalodata={row.get('kalodata_url')}")
        return 0

    repository = None if args.no_rds else CreatorRepository.from_env(args.creator_db_url)
    asset_storage = CreatorAssetStorage()
    video_fetcher = VideoFetcherAgent()
    grid_generator = GridGeneratorAgent()
    scoring_agent = CombinedScoringVibeAgent()
    category_agent = CategoryTaggingAgent()
    entry_agent = EntryScreenAgent()

    result_path = OUT_DIR / f"excel_outreach_{args.batch_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jsonl"
    processed: List[Dict[str, Any]] = []
    failed = 0
    stopped_reason = ""
    with result_path.open("w", encoding="utf-8") as output:
        for index, row in enumerate(rows, 1):
            started = time.time()
            print(f"\n[{index}/{len(rows)}] @{row['tk_handle']}")
            try:
                result = process_one(
                    row,
                    repository=repository,
                    asset_storage=asset_storage,
                    video_fetcher=video_fetcher,
                    grid_generator=grid_generator,
                    scoring_agent=scoring_agent,
                    category_agent=category_agent,
                    entry_agent=entry_agent,
                    max_videos=args.max_videos,
                    max_grids=args.max_grids,
                    retention_days=args.retention_days,
                )
                result["elapsed_sec"] = round(time.time() - started, 1)
                processed.append(result)
                if result.get("skipped"):
                    print(f"  ↪️ 跳过: {result.get('skip_reason')}")
                else:
                    entry = result.get("entry_result") or {}
                    print(f"  ✅ {entry.get('entry_decision')} score={entry.get('entry_score')}")
            except Exception as exc:
                failed += 1
                row["error"] = str(exc)
                row["elapsed_sec"] = round(time.time() - started, 1)
                processed.append(row)
                print(f"  ❌ 失败: {exc}")
                if isinstance(exc, KalodataAccessBlocked) and not args.continue_on_kalodata_block:
                    stopped_reason = "Kalodata 403 风控熔断：已停止整批，避免继续刷失败请求"
                    row["stopped_reason"] = stopped_reason
                    print(f"  🛑 {stopped_reason}")
            output.write(json.dumps(processed[-1], ensure_ascii=False, default=str) + "\n")
            output.flush()
            if stopped_reason:
                break

    passed = [
        row for row in processed
        if not row.get("skipped")
        and not row.get("error")
        and (row.get("entry_result") or {}).get("entry_decision") == "通过"
    ]
    if args.include_manual:
        passed.extend([
            row for row in processed
            if not row.get("skipped")
            and not row.get("error")
            and (row.get("entry_result") or {}).get("entry_decision") == "人工查看"
        ])

    print(f"\n📌 最终建联名单: {len(passed)}；失败: {failed}；明细: {result_path}")
    feishu_created = 0
    if passed:
        batch_message = generate_batch_outreach(
            passed,
            product_name=args.product_name,
            product_category=args.product_category,
            selling_points=args.selling_points,
            version=args.message_version,
        )
        output_rows = [
            build_output_row(
                row,
                batch_id=args.batch_id,
                product_name=args.product_name,
                batch_message=batch_message,
                include_oss_links=args.include_oss_links,
            )
            for row in passed
        ]
        final_path = OUT_DIR / f"final_outreach_{args.batch_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        final_path.write_text(json.dumps(output_rows, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        print(f"📝 最终输出 JSON: {final_path}")

        if args.write_feishu:
            if not args.output_feishu_url:
                raise RuntimeError("--write-feishu 需要传 --output-feishu-url")
            app_token, table_id = parse_feishu_url(args.output_feishu_url)
            ensure_final_fields(app_token, table_id)
            feishu_created = create_feishu_records(app_token, table_id, output_rows)
            print(f"✅ 已写入最终飞书表: {feishu_created} 条")

    print(json.dumps({
        "excel_count": len(rows),
        "processed": len(processed),
        "final_outreach_count": len(passed),
        "failed": failed,
        "stopped_reason": stopped_reason,
        "feishu_created": feishu_created,
        "result_path": str(result_path),
    }, ensure_ascii=False, indent=2))
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
